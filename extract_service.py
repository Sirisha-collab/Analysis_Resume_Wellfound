import os
import re
import pdfplumber
import pandas as pd
import spacy

from docx import Document
from datetime import datetime
from dateutil import parser

from rapidfuzz import fuzz
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

from constants import SDR_SKILLS

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# =====================================================
# CONFIG
# =====================================================

RESUME_FOLDER = r"D:\Docs Latest\A Masters Required doc\Projects\Resumes Wellfound\Resumes SDR"
JOB_DESCRIPTION_FILE = r"D:\Docs Latest\A Masters Required doc\Projects\Resumes Wellfound\Resumes SDR\job description.txt"

nlp = spacy.load("en_core_web_lg")

embedding_model = SentenceTransformer(
    "sentence-transformers/all-MiniLM-L6-v2"
)

# =====================================================
# EXTRACTORS
# =====================================================

def extract_name(text):
    doc = nlp(text[:3000])

    for ent in doc.ents:
        if ent.label_ == "PERSON":
            return ent.text

    return ""


def extract_email(text):
    pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b"

    matches = re.findall(pattern, text)

    return matches[0].lower() if matches else ""


def extract_phone(text):
    candidates = re.findall(
        r'[\+]?\d[\d\s\-\(\)]{8,20}\d',
        text
    )

    for candidate in candidates:

        digits = re.sub(r"\D", "", candidate)

        if len(digits) == 12 and digits.startswith("91"):
            digits = digits[2:]

        if len(digits) == 10:
            return digits

    return ""


def extract_linkedin(text):
    pattern = r"(https?://(?:www\.)?linkedin\.com/[^\s]+|linkedin\.com/[^\s]+)"

    match = re.search(
        pattern,
        text,
        re.IGNORECASE
    )

    return match.group(1) if match else ""


def extract_github(text):
    pattern = r"(https?://(?:www\.)?github\.com/[^\s]+|github\.com/[^\s]+)"

    match = re.search(
        pattern,
        text,
        re.IGNORECASE
    )

    return match.group(1) if match else ""


def extract_location(text):
    doc = nlp(text[:2000])

    locations = []

    for ent in doc.ents:
        if ent.label_ in ["GPE", "LOC"]:
            locations.append(ent.text)

    locations = list(dict.fromkeys(locations))

    return ", ".join(locations[:3])


def extract_companies(text):
    doc = nlp(text)

    companies = []

    for ent in doc.ents:
        if ent.label_ == "ORG":
            companies.append(ent.text)

    companies = list(dict.fromkeys(companies))

    return ", ".join(companies[:15])


# =====================================================
# EXPERIENCE
# =====================================================
def parse_date(date_str):

    date_str = date_str.lower().strip()

    if "present" in date_str or "current" in date_str:
        return datetime.today()

    try:
        return parser.parse(
            date_str,
            fuzzy=True
        )
    except:
        return None


def extract_date_ranges(text):

    pattern = (
        r'([A-Za-z]{3,9}\s*\d{4}|\d{4})'
        r'\s*[-–to]+\s*'
        r'([A-Za-z]{3,9}\s*\d{4}|\d{4}|present|current)'
    )

    return re.findall(
        pattern,
        text,
        re.IGNORECASE
    )


def compute_experience(text):

    ranges = extract_date_ranges(text)

    if not ranges:
        return 0

    total_days = 0

    for start, end in ranges:

        start_date = parse_date(start)
        end_date = parse_date(end)

        if not start_date or not end_date:
            continue

        if end_date < start_date:
            continue

        total_days += (
            end_date - start_date
        ).days

    return round(
        total_days / 365.25,
        1
    )


# =====================================================
# SKILLS
# =====================================================

def extract_skills(text):

    text = text.lower()

    found = []

    for skill in SDR_SKILLS:

        score = fuzz.partial_ratio(
            skill,
            text
        )

        if score >= 85:
            found.append(skill)

    return sorted(list(set(found)))


# =====================================================
# SEMANTIC MATCHING
# =====================================================

def create_embedding(text):

    return embedding_model.encode(
        text,
        normalize_embeddings=True
    )


def semantic_similarity(
    resume_text,
    job_embedding
):

    resume_embedding = create_embedding(
        resume_text
    )

    similarity = cosine_similarity(
        [resume_embedding],
        [job_embedding]
    )[0][0]

    return float(similarity)


# =====================================================
# SCORING
# =====================================================

def score_candidate(candidate, job_embedding):

    resume_embedding = candidate["Embedding"]

    # 1. Semantic similarity (MAIN SIGNAL)
    semantic_score = cosine_similarity(
        [resume_embedding],
        [job_embedding]
    )[0][0]

    score = semantic_score * 100
    reasons = [f"Semantic match ({semantic_score:.2f})"]

    # 2. Experience bonus
    exp = float(candidate.get("Experience (Years)", 0) or 0)

    if exp >= 1:
        score += min(exp * 2, 20)
        reasons.append(f"Experience bonus ({exp})")

    # 3. Skills signal
    skills = candidate.get("Skills", "")
    skill_count = len(skills.split(",")) if skills else 0

    score += min(skill_count * 2, 20)
    reasons.append(f"Skills count ({skill_count})")

    # 4. LinkedIn boost (important in hiring pipelines)
    if candidate.get("LinkedIn"):
        score += 10
        reasons.append("LinkedIn present (+10)")

    return round(score, 2), "; ".join(reasons)

# =====================================================
# EXCEL STYLING
# =====================================================

def style_excel(file_path):

    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_fill = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center"
        )

    for col in ws.columns:

        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(
                    max_len,
                    len(str(cell.value))
                )

        ws.column_dimensions[
            col_letter
        ].width = max_len + 5

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)


# =====================================================
# MAIN
# =====================================================

def main():

    with open(
        JOB_DESCRIPTION_FILE,
        "r",
        encoding="utf-8"
    ) as f:
        job_description = f.read()

    job_embedding = create_embedding(
        job_description
    )

    results = []

    files = os.listdir(
        RESUME_FOLDER
    )

    print(
        f"Files Found: {len(files)}"
    )

    for file in files:

        path = os.path.join(
            RESUME_FOLDER,
            file
        )

        if file.lower().endswith(".pdf"):
            text = read_pdf(path)

        elif file.lower().endswith(".docx"):
            text = read_docx(path)

        else:
            continue

        if not text.strip():
            continue

        skills = extract_skills(text)

        results.append({

            "File": file,
            "Name": extract_name(text),
            "Email": extract_email(text),
            "Phone": extract_phone(text),
            "Location": extract_location(text),
            "Companies": extract_companies(text),
            "Experience (Years)": compute_experience(text),
            "Skills": ", ".join(skills),
            "LinkedIn": extract_linkedin(text),
            "GitHub": extract_github(text),
            "RawText": text
        })

    # Candidate Repo
    candidate_df = pd.DataFrame(results)

    timestamp = datetime.now().strftime( "%Y%m%d_%H%M%S" )

    repo_file = (f"candidate_repo_{timestamp}.xlsx") 

    candidate_df.to_excel(repo_file, index=False)

    style_excel(repo_file)
    print(f"Candidate Repo Saved: {repo_file}")

    # Ranking
    ranked = []
    for r in results:

        score, reason = score_candidate(r,job_embedding)

        if score < 0:
            continue

        r["Score"] = score
        r["Reason"] = reason

        ranked.append(r)

    top5 = sorted(
        ranked,
        key=lambda x: x["Score"],
        reverse=True
    )[:5]

    top5_df = pd.DataFrame(top5)

    top5_file = (f"SDR_top_5_{timestamp}.xlsx")

    top5_df.to_excel(top5_file, index=False)

    style_excel(top5_file)

    print(f"Top 5 Saved: {top5_file}")


if __name__ == "__main__":
    main()