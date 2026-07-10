import os
import re
import pandas as pd
from datetime import datetime

from ingestion_service import read_pdf, read_docx
from extract_service import (
    extract_name, extract_email, extract_phone,
    extract_location, compute_experience,
    extract_skills, extract_linkedin, extract_github,
    create_embedding
)

from sklearn.metrics.pairwise import cosine_similarity
import spacy

# -----------------------------
# SPEED OPTIMIZATION
# -----------------------------
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

nlp = spacy.load("en_core_web_lg")

# -----------------------------
# CONFIG
# -----------------------------
RESUME_FOLDER = r"D:\Docs Latest\A Masters Required doc\Projects\Resumes Wellfound\Resumes SDR"
JOB_DESCRIPTION_FILE = r"D:\Docs Latest\A Masters Required doc\Projects\Resumes Wellfound\Resumes SDR\job description.txt"

# -----------------------------
# LOAD JD + EMBEDDING
# -----------------------------
with open(JOB_DESCRIPTION_FILE, "r", encoding="utf-8") as f:
    job_description = f.read()

job_embedding = create_embedding(job_description)

def clean_text(value):
    if not isinstance(value, str):
        return value
    return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', value)

def safe_exp(val):
    try:
        return float(val)
    except:
        return 0.0

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# EXCEL STYLING

def style_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 5

    wb.save(file_path)


# -----------------------------
# SCORING FUNCTION
# -----------------------------
def score_candidate(r, job_embedding):

    score = 0
    reasons = []

    skills_raw = r.get("Skills") or ""

    if isinstance(skills_raw, list):
        skills_text = ", ".join(skills_raw).lower()
    else:
        skills_text = str(skills_raw).lower()

    full_text = str(r.get("RawText") or "").lower()

    # EXPERIENCE
    exp = safe_exp(r.get("Experience (Years)", 0))

    if exp < 1:
        return -1, "Rejected: less than 1 year experience"

    score += 10 if exp <= 8 else 5

    # -----------------------------
    # SEMANTIC MATCH
    # -----------------------------
    resume_embedding = create_embedding(full_text)

    sim = cosine_similarity(
        [resume_embedding],
        [job_embedding]
    )[0][0]

    score += sim * 100
    reasons.append(f"JD similarity ({sim:.2f})")

    # -----------------------------
    # SIGNALS
    # -----------------------------
    if r.get("LinkedIn"):
        score += 10
        reasons.append("LinkedIn (+10)")

    if r.get("GitHub"):
        score += 15
        reasons.append("GitHub (+15)")

    # -----------------------------
    # NLP SIGNAL
    # -----------------------------
    doc = nlp(full_text[:2000])

    if any(ent.label_ == "ORG" for ent in doc.ents):
        score += 5
        reasons.append("Company experience (+5)")

    return round(score, 2), "; ".join(reasons)

# -----------------------------
# MAIN PIPELINE
# -----------------------------
all_candidates = []
ranked_candidates = []

files = os.listdir(RESUME_FOLDER)
print(f"Total files found: {len(files)}")

for file in files:

    if file.startswith("."):
        continue

    if not file.lower().endswith((".pdf", ".docx")):
        continue

    path = os.path.join(RESUME_FOLDER, file)

    try:
        if file.lower().endswith(".pdf"):
            text = read_pdf(path)
        else:
            text = read_docx(path)
    except:
        continue

    if not text or len(text.strip()) < 50:
        continue

    exp = safe_exp(compute_experience(text))

    row = {
        "File": file,
        "Name": extract_name(text),
        "Email": extract_email(text),
        "Phone": extract_phone(text),
        "Location": extract_location(text),
        "Experience (Years)": exp,
        "Skills": extract_skills(text),
        "LinkedIn": extract_linkedin(text),
        "GitHub": extract_github(text),
        "RawText": text
    }

    all_candidates.append(row)

    # FILTER
    if exp < 1:
        continue

    score, reason = score_candidate(row, job_embedding)

    if score < 0:
        continue

    row["Score"] = score
    row["Score Breakdown"] = reason

    ranked_candidates.append(row)


def safe_dataframe(df):
    df = df.copy()

    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: clean_text(x)
            if isinstance(x, str)
            else ", ".join(x) if isinstance(x, list)
            else str(x) if x is not None
            else ""
        )

    return df


SDR_DIR = "SDR_results"

os.makedirs(SDR_DIR, exist_ok=True)

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

# candidate repo -----------------------------
df_all = pd.DataFrame(all_candidates)
df_all = safe_dataframe(df_all)

repo_file =  os.path.join(
    SDR_DIR,
    f"sdr_candidate_repo_{timestamp}.xlsx"
)

df_all.to_excel(repo_file, index=False)

style_excel(repo_file)

print("\n Candidate Repo created:", repo_file)
print("Total resumes parsed:", len(df_all))


# top 5 -----------------------------
top5_sorted = sorted(
    ranked_candidates,
    key=lambda x: x.get("Score", 0),
    reverse=True
)[:5]

df_top5 = pd.DataFrame(top5_sorted)
df_top5 = safe_dataframe(df_top5)

top5_file =  os.path.join(
    SDR_DIR,
    f"sdr_top_5_{timestamp}.xlsx"
)
df_top5.to_excel(top5_file, index=False)

style_excel(top5_file)

print("\n🏆 Top 5 file created:", top5_file)
print("Qualified candidates:", len(ranked_candidates))