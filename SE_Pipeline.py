import os
from pydoc import text
import re
import pandas as pd
import requests
from datetime import datetime

from ingestion_service import read_pdf, read_docx, read_csv
from extract_service import (
    extract_name, extract_email, extract_phone,
    extract_location, compute_experience,
     extract_linkedin, extract_github,create_embedding
)      

from rapidfuzz import fuzz
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity

import spacy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# -----------------------------
# EXCEL STYLING
# -----------------------------
def style_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 5

    wb.save(file_path)


# -----------------------------
# CONFIG
# -----------------------------
RESUME_FOLDER = r"D:\Docs Latest\A Masters Required doc\Projects\Resumes Wellfound\Resume SE"
JOB_DESCRIPTION_FILE = r"D:\Docs Latest\A Masters Required doc\Projects\Resumes Wellfound\Resume SE\job description.txt"
CSV_FILE = r"D:\Docs Latest\A Masters Required doc\Projects\Resumes Wellfound\Resume SE\Applicants.csv"

with open(JOB_DESCRIPTION_FILE, "r", encoding="utf-8") as f:
    job_description = f.read()

job_embedding = create_embedding(job_description)

SE_SKILLS = [
"full stack development",
"react",
"node.js",
"sql",
"Typescript",
"rest api development",
"AWS",
"saas integrations",
"Azure",
"agile methodologies",
"postgresql",
"ci/cd pipelines",
"devops",
"cloud",
"express.js",
"sql",
"linux",
"integration testing",
"code review",
"git version control",
"performance optimization",
"Redis"
]

#exp into 3.5 instead of 3,
def safe_exp(val):
    try:
        return float(val)
    except:
        return 0.0

def extract_skills(text):

    text = text.lower()

    found = []

    for skill in SE_SKILLS:

        score = fuzz.partial_ratio(
            skill,
            text
        )

        if score >= 85:
            found.append(skill)

    return sorted(list(set(found)))

# -----------------------------
# METRICS DETECTION
# -----------------------------
def extract_metrics(row):
    text = " ".join([
        str(row.get("name", "")),
        str(row.get("location_city", "")),
        str(row.get("canonical_skills", ""))
    ]).lower()

    patterns = [
        r"\d+\+?\s*%",
        r"improved", r"built", r"developed",
        r"deployed", r"scaled", r"optimized",
        r"reduced", r"increased", r"architected"
    ]

    return any(re.search(p, text) for p in patterns)


#GITHUB QUALITY EVALUATION
# def evaluate_github_quality(r):
#     score = 0
#     reasons = []

#     repos = r.get("github_repos") or []
#     commits = r.get("github_commit_count", 0)

#     if not repos:
#         return -30, ["No GitHub repositories found"]

#     original_repos = [
#         repo for repo in repos
#         if not repo.get("fork", False)
#     ]

#     fork_ratio = 1 - (len(original_repos) / len(repos))

#     if fork_ratio > 0.70:
#         score -= 40
#         reasons.append("Mostly forked repositories (-40)")
#     else:
#         score += 10
#         reasons.append("Original repositories (+10)")

#     if commits < 20:
#         score -= 40
#         reasons.append("Low GitHub activity (-40)")

#     elif commits >= 100:
#         score += 20
#         reasons.append("Strong GitHub activity (+20)")

#     else:
#         score += 10
#         reasons.append("Some GitHub activity (+10)")

#     copy_words = [
#         "tutorial",
#         "course",
#         "clone",
#         "starter",
#         "boilerplate",
#         "assignment",
#         "udemy",
#         "youtube"
#     ]

#     copy_projects = 0

#     for repo in repos:
#         text = (
#             str(repo.get("name", "")) +
#             str(repo.get("description", "")) +
#             str(repo.get("readme", ""))
#         ).lower()

#         if any(word in text for word in copy_words):
#             copy_projects += 1


#     if copy_projects >= 2:
#         score -= 30
#         reasons.append("Possible tutorial/copy projects (-30)")

#     engineering_terms = [
#         "docker",
#         "docker-compose",
#         "redis",
#         "postgres",
#         "postgresql",
#         "mongodb",
#         "authentication",
#         "oauth",
#         "jwt",
#         "testing",
#         "pytest",
#         "jest",
#         "github actions",
#         "ci/cd",
#         "kubernetes",
#         "queue",
#         "microservice",
#         "api"
#     ]

#     github_text = " ".join(
#         [
#             str(repo.get("description", "")) +
#             str(repo.get("readme", ""))
#             for repo in repos
#         ]
#     ).lower()


#     depth_matches = sum(
#         1 for term in engineering_terms
#         if term in github_text
#     )


#     if depth_matches >= 5:
#         score += 20
#         reasons.append("Good engineering depth (+20)")

#     elif depth_matches == 0:
#         score -= 15
#         reasons.append("No technical depth found (-15)")


#     return score, reasons

# -----------------------------
# SOFTWARE ENGINEER SCORING
# -----------------------------
def score_candidate(r):
    score = 0
    reasons = []
    negative_penalty = 0

    skills_raw = r.get("canonical_skills") or ""

    if isinstance(skills_raw, list):
        skills_text = " ".join(skills_raw).lower()
    else:
        skills_text = str(skills_raw).lower()

    exp = safe_exp(r.get("years_experience_in_role", 0))

    linkedin = r.get("linkedin_url")
    github = r.get("github_url")

    # rejection rules
    if exp < 2:
        return -1, "Rejected: less than 2 years experience",""

    if not linkedin or not github:
        return -1, "Rejected: missing LinkedIn or GitHub", ""

    if 2 <= exp <= 8:
        score += 15
    elif exp > 8:
        score += 5

    HIGH = [
        "nodejs", "node.js",
        "react",
        "typescript",
        "javascript",
        "express",
        "nextjs",
        "next.js",
        "aws",
        "redis"
    ]

    MEDIUM = [
        "postgresql",
        "sql",
        "graphql",
        "rest api",
        "restful api",
        "git",
        "system design"
    ]

    PENALTIES = {
    "ai": -40,
    "artificial intelligence": -40,
    "blockchain": -40,
    "html": -40,
    "css": -40,
    "docker": -40,
    "java": -40,
    ".net": -40,
    "kubernetes": -40,
    "bootstrap": -30,
    "php": -20
}
    
    REQUIRED_BACKEND = [
        "nodejs",
        "node.js",
        "python",
        "java",
        "golang",
        "express"
    ]

    REQUIRED_FRONTEND = [
        "react",
        "javascript",
        "typescript"
    ]

    NEGATIVE = [
        "sales",
        "marketing",
        "sdr",
        "cold calling",
        "lead generation",
        "public relations",
        "seo"
    ]
    negative_skills =[]

    has_backend = any(x in skills_text for x in REQUIRED_BACKEND)
    has_frontend = any(x in skills_text for x in REQUIRED_FRONTEND)

    if not has_backend:
        return -1, "Rejected: no backend engineering skills", ""

    if not has_frontend:
        return -1, "Rejected: no frontend engineering skills", ""

    # Skill scoring
    matched = set()

    for skill in HIGH:
        if skill in skills_text and skill not in matched:
            score += 20
            reasons.append(f"{skill}")
            matched.add(skill)

    for skill in MEDIUM:
        if skill in skills_text and skill not in matched:
            score += 10
            reasons.append(f"{skill} ")
            matched.add(skill)

    # Full stack requirement
    if has_frontend and has_backend:
        score += 30
        reasons.append("Full-stack capability")

    # Cloud / deployment bonus
    if any(x in skills_text for x in ["aws", "azure", "gcp"]):
        score += 10
        reasons.append("Cloud experience")

    if "git" in skills_text:
        score += 10
        reasons.append("Git experience")

    # Evidence of engineering work
    if extract_metrics(r):
        score += 20
        reasons.append(" ")

    # Profile signals
    score += 10
    reasons.append(" ")

    # Apply skill penalties
    for skill, penalty in PENALTIES.items():
        if skill in skills_text:
            score += penalty
            reasons.append(f"{skill} ({penalty})")
            negative_skills.append(f"{skill}: {penalty}")

    # Apply non-engineering penalties
    for skill in NEGATIVE:
        if skill in skills_text:
            score -= 70
            reasons.append(f"{skill} (-70)")
            negative_skills.append(f"{skill}: -70")

    # Final quality gate
    if score < 70:
        negative_reason = ""
        if negative_skills:
            negative_reason = " | Negative skills: " + ", ".join(negative_skills)
    
    return score, "; ".join(reasons), ", ".join(negative_skills)

#######################
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# import time

# options = webdriver.ChromeOptions()
# options.add_argument(r"user-data-dir=C:\selenium_profile")
# options.add_argument("--profile-directory=Default")
# options.add_argument("--no-sandbox")
# options.add_argument("--disable-blink-features=AutomationControlled")
# options.add_experimental_option("excludeSwitches", ["enable-automation"])
# options.add_experimental_option('useAutomationExtension', False)

# driver = webdriver.Chrome(options=options)

# url = "https://wellfound.com/link/15163639/615eab389771dd8d67c5495980563961/github_url"

# try:
#     driver.get(url)
#     time.sleep(3)
    
#     # Check if access is restricted
#     page_source = driver.page_source
#     if "access restricted" in page_source.lower():
#         print("❌ Access Restricted - Session not valid")
#         print("Try: Open Chrome manually, go to that Wellfound link and copy redirected URL")
#         # Print page content to debug
#         print(driver.current_url)
#         print(page_source[:500])
#     else:
#         # Wait for redirect
#         WebDriverWait(driver, 10).until(lambda d: d.current_url != url)
#         github_url = driver.current_url
#         print(github_url)
        
# finally:
#     driver.quit()

# -----------------------------
# CSV ONLY PIPELINE
# -----------------------------

all_candidates = []
ranked_candidates = []

print("Reading CSV:", CSV_FILE)


try:
    df = pd.read_csv(
        CSV_FILE,
        encoding="utf-8",
        sep=",",
        engine="python"
    )
except UnicodeDecodeError:
    print("UTF-8 failed, trying cp1252...")
    df = pd.read_csv(
        CSV_FILE,
        encoding="cp1252",
        sep=",",
        engine="python"
    )

df.columns = df.columns.str.strip()

print("CSV loaded successfully")
print("Total applicants:", len(df))

for _, row_data in df.iterrows():

    row = {
        "name": row_data.get("name", ""),
        "Email": row_data.get("Email", ""),
        "Phone": row_data.get("Phone", ""),
        "location_city": row_data.get("location_city", ""),
        "years_experience_in_role": safe_exp(
            row_data.get("years_experience_in_role", 0)
        ),
        "canonical_skills": row_data.get("canonical_skills", ""),
        "linkedin_url": row_data.get("linkedin_url", ""),
        "github_url": row_data.get("github_url", "")
    }

    all_candidates.append(row)

    score, reason, negative_penalty = score_candidate(row)

    if score < 0:
        continue

    row["Score"] = score
    row["Score Breakdown"] = reason
    row["Negative Penalty"] = negative_penalty

    ranked_candidates.append(row)


# -----------------------------
# EXPORT RESULTS
# -----------------------------

SE_DIR = "SE_results"
os.makedirs(SE_DIR, exist_ok=True)

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')


# Complete CSV repository
df_all = pd.DataFrame(all_candidates)

repo_file = os.path.join(
    SE_DIR,
    f"se_candidate_repo_{timestamp}.xlsx"
)

df_all.to_excel(repo_file, index=False)
style_excel(repo_file)

print("Candidate repo saved:", repo_file)


# Top 50 candidates
top50 = sorted(
    ranked_candidates,
    key=lambda x: x["Score"],
    reverse=True
)[:50]


df_top50 = pd.DataFrame(top50)

top50_file = os.path.join(
    SE_DIR,
    f"se_top_50_{timestamp}.xlsx"
)

df_top50.to_excel(top50_file, index=False)
style_excel(top50_file)

print("Top 50 saved:", top50_file)

# # -----------------------------
# # PIPELINE
# # -----------------------------
# all_candidates = []
# ranked_candidates = []

# files = os.listdir(RESUME_FOLDER)
# print(f"Total files found: {len(files)}")

# for file in files:

#     if file.startswith("."):
#         continue

#     if not file.lower().endswith((".pdf", ".docx", ".csv")):
#         continue

#     path = os.path.join(RESUME_FOLDER, file)

#     try:
#         if file.lower().endswith(".pdf"):
#             text = read_pdf(path)
#         elif file.lower().endswith(".docx"):
#             text = read_docx(path)
#         elif file.lower().endswith(".csv"):
#             text = read_csv(path)
#         else:
#             continue
#     except Exception as e:
#         print(f"Error reading {file}: {e}")
#         continue

#     if not text or len(text.strip()) < 50:
#         continue

#     exp = safe_exp(compute_years_experience_in_role(text))

#     row = {
#         "File": file,
#         "name": extract_name(text),
#         "Email": extract_email(text),
#         "Phone": extract_phone(text),
#         "location_city": extract_location(text),
#         "years_experience_in_role (Years)": exp,
#         "Skills": extract_skills(text),
#         "LinkedIn": extract_linkedin(text),
#         "GitHub": extract_github(text)
#     }

#     all_candidates.append(row)

#     if exp < 1:
#         continue

#     score, reason = score_candidate(row)

#     if score < 0:
#         continue

#     row["Score"] = score
#     row["Score Breakdown"] = reason

#     ranked_candidates.append(row)


# -----------------------------
# EXPORT
# -----------------------------

SE_DIR = "SE_results"
os.makedirs(SE_DIR, exist_ok=True)

timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

df_all = pd.DataFrame(all_candidates)

repo_file =  os.path.join(
    SE_DIR,
    f"se_candidate_repo_{timestamp}.xlsx"
)
df_all.to_excel(repo_file, index=False)
style_excel(repo_file)

print(" SE Candidate repo saved:", repo_file)


top5 = sorted(ranked_candidates, key=lambda x: x["Score"], reverse=True)[:5]

df_top5 = pd.DataFrame(top5)

top5_file =  os.path.join(
    SE_DIR,
    f"se_top_5_{timestamp}.xlsx"
)
df_top5.to_excel(top5_file, index=False)
style_excel(top5_file)

print(" SE Top 5 saved:", top5_file)